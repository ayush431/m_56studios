from .serializer import UploadSerializer
from django.core.files.storage import FileSystemStorage
import os
import tempfile
import time
import openpyxl
import datetime
from datetime import datetime
from common.constant import INVALID_REQUEST_TRY_AGAIN
from app.question.serializer import QuestionSerializer
from app.options.serializer import OptionsSerializer
from app.question.controller import QuestionController
from app.options.controller import OptionsController
from common.utility import isValidExcelFile
from app.question.model import question
from app.status.model import status_log
from app.utility.db_utility import performBatchOperation
from django.db import transaction

class FileUploadController():
    def save_data(self, vendor_id, vendor_name, obj, user):
        if obj.is_valid(raise_exception=True):
            fs = FileSystemStorage(tempfile.gettempdir())
            file_name = fs.save(content=obj.validated_data['file'], name= str(time.time()) + ".xlsx")
            full_path = os.path.join(fs.location, file_name)
            
            workbook = openpyxl.load_workbook(full_path)
            sheet = workbook.active
            first_row = [] # The row where we stock the name of the column
            for col in range(1, sheet.max_column+1):
                first_row.append(sheet.cell(row=1, column=col).value)
            data =[]
            for row in range(2, sheet.max_row+1):
                elm = {}
                for col in range(1, sheet.max_column+1):
                    if first_row[col-1] in elm:
                        temp = elm[first_row[col-1]]
                        if type(temp) != list:
                            elm[first_row[col-1]] = []
                            elm[first_row[col-1]].append(temp)
                        elm[first_row[col-1]].append(sheet.cell(row=row,column=col).value)
                    else:
                        elm[first_row[col-1]]=sheet.cell(row=row,column=col).value
                data.append(elm)
            
            # Check fields in excel sheet
            if not isValidExcelFile(data):
                raise Exception("Invalid Excel file, fields are missing")
            allValidQuestions = []
            for each_question in data:
                try: 
                    if not each_question["QUESTION"]:
                        continue

                    # Check if category is valid
                    if each_question["CATEGORY"] not in question.categories:
                        raise Exception("Invalid category found {}".format(each_question["CATEGORY"]))

                    # Check if difficulty level is valid
                    if each_question["DIFFICULT"] not in question.difficult_level:
                        raise Exception("Invalid difficulty level found {}".format(each_question["DIFFICULT"]))

                    all_options = []
                    rightanswer_found = False
                    for each_option in each_question["CHOICE"]:
                        
                        if each_option == None or len(str(each_option)) < 1:
                            continue

                        if each_option == each_question["RIGHT_ANSWER"]:
                            rightanswer_found = True

                        option_obj = {}
                        option_obj["option"] = each_option
                        option_obj["is_correct"] = True if each_option == each_question["RIGHT_ANSWER"]  else False
                        all_options.append(option_obj)

                    if not rightanswer_found:
                        raise Exception("Right answer not found for question  {}".format(each_question["QUESTION"]))

                    if len(all_options) < 3 or len(all_options) >= 6:
                        if each_question["CATEGORY"] == question.categories.TRUE_OR_FALSE:
                            if len(all_options) < 2:
                                raise Exception("Invalid number of options found for question  {}".format(each_question["QUESTION"]))
                        else:
                            raise Exception("Invalid number of options found for question  {}".format(each_question["QUESTION"]))

                    questionObj = {
                        "question": each_question["QUESTION"],
                        "qs_date": datetime.now(),
                        "category": each_question["CATEGORY"],
                        "is_image_hint": each_question["ISIMAGEAHINT"],
                        "difficulty_level": each_question["DIFFICULT"],
                        "user_id":str(user.user_id),
                        "cropped_json": "{}",
                        "system_check": "Valid",
                        "original_qn_id": each_question["ID"],
                        "rejected_by_editor": False,
                        "should_overwrite": False,
                        "vendor_id": vendor_id,
                        "vendor_name": vendor_name,
                        "status": "Draft",
                        "discarded_by_editor": False,
                        "resrc_type": each_question["RESRCTYPE"] if "RESRCTYPE" in each_question else None,
                        "reveal_image_url": each_question["REVEALIMAGEURL"] if "REVEALIMAGEURL" in each_question else None,
                        "resrc_url": each_question["RESRCURL"] if "RESRCURL" in each_question else None,
                        "original_url": each_question["RESRCURL"] if "RESRCURL" in each_question else None,
                        "offline_content": each_question["OFFLINE"] if "OFFLINE" in each_question else False,
                    }

                    if user.user_type == "reviewer":
                        questionObj["status"] = "For Review"
                        questionObj["qn_submitted_date"] = datetime.now().date()
                    
                    # questionObj = QuestionController().question_check(questionObj, all_options, user.user_type)
                    questionObj["CHOICE"] = all_options
                    allValidQuestions.append(questionObj)
                except Exception as e:
                    raise Exception(e)
                # Add options and then return


            status_logs = []
            all_options = []
            # Start inserting data if everything is valid
            for questionObj in allValidQuestions:
                questionObj = QuestionController().question_check(questionObj, questionObj["CHOICE"], user.user_type)
                qn_instance = QuestionController().save_data(obj=QuestionSerializer(data=questionObj))
                status_logs.append(status_log(old_status=None, new_status=questionObj["status"], qn_id_id=qn_instance.get('qn_id'), user_id=user))
                for each_option in questionObj["CHOICE"]:
                    if not each_option or len(str(each_option)) < 1:
                        continue
                    each_option["qn_id"] = qn_instance.get('qn_id')
                    each_option["option"] = str(each_option["option"])
                    # all_options.append(options(option=str(each_option["option"]), is_correct=each_option["is_correct"], qn_id_id=qn_instance.get('qn_id')))
                    OptionsController().save_data(obj=OptionsSerializer(data=each_option))
            performBatchOperation(status_log, status_logs, 'create')
            return True
        return False
        
    def update_flag_qus(self, request, obj, user):
          if obj.is_valid(raise_exception=True):
            fs = FileSystemStorage(tempfile.gettempdir())
            file_name = fs.save(content=obj.validated_data['file'], name= str(time.time()) + ".xlsx")
            full_path = os.path.join(fs.location, file_name)
            
            workbook = openpyxl.load_workbook(full_path)
            sheet = workbook.active
            first_row = [] # The row where we stock the name of the column
            for col in range(1, sheet.max_column+1):
                first_row.append(sheet.cell(row=1, column=col).value)
            data =[]
            for row in range(2, sheet.max_row+1):
                elm = {}
                for col in range(1, sheet.max_column+1):
                    if first_row[col-1] in elm:
                        temp = elm[first_row[col-1]]
                        if type(temp) != list:
                            elm[first_row[col-1]] = []
                            elm[first_row[col-1]].append(temp)
                        elm[first_row[col-1]].append(sheet.cell(row=row,column=col).value)
                    else:
                        elm[first_row[col-1]]=sheet.cell(row=row,column=col).value
                data.append(elm)
            qus_array = []
            qn_id_array = []
            status_update_logs = []
            with transaction.atomic():
                for each_data in data:
                    if each_data["QUESTIONID"] == None:
                        break
                    qus_data, boolean_value = QuestionController().get_qus_id_object(each_data["QUESTIONID"])
                    if not boolean_value:
                        qn_id_array.append(qus_data)
                    if boolean_value:
                        if qus_data and each_data["STATUS"] != qus_data["status"]:
                            qus_array.append(question(status=each_data["STATUS"], flagged_reason=each_data["REASONS"], qn_id=each_data["QUESTIONID"]))
                            status_update_logs.append(status_log(old_status=qus_data["status"], new_status=each_data["STATUS"], qn_id_id=each_data["QUESTIONID"], user_id=user))
                if qn_id_array != []:
                    raise Exception("Invalid question IDs Supplied - Total {} QuestionIDs {}".format(len(qn_id_array), qn_id_array))
                # Bulk update questions
                performBatchOperation(question, qus_array, 'update', fields=['status', 'flagged_reason'])
                # Bulk create status logs
                performBatchOperation(status_log, status_update_logs, 'create')


    def post(self, vendor_id, vendor_name, request=None):
        if self.check_request(request=request):
            if request.user.user_type == "manager":
                return self.update_flag_qus(request, obj=UploadSerializer(data=request.data), user=request.user)
            return self.save_data(vendor_id, vendor_name, obj=UploadSerializer(data=request.data), user=request.user)

    def check_request(self, request):
        if request and request.data:
            return True
        raise Exception(INVALID_REQUEST_TRY_AGAIN)
